from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as exc:
    raise ImportError(
        "openpyxl is required to parse the GatorBio Excel form. "
        "Install it with: pip install openpyxl"
    ) from exc

from .common import (
    SampleType,
    map_assay_label_to_code,
    map_max_plate_label_to_code,
)


def read_excel_file(excel_path: Path) -> Dict[str, "Worksheet"]:
    """Read and parse the Excel file with multiple sheets (supports .xlsx and .xlsm)."""

    workbook = load_workbook(excel_path, data_only=True)
    sheets: Dict[str, "Worksheet"] = {}
    for sheet_name in workbook.sheetnames:
        sheets[sheet_name] = workbook[sheet_name]

    print(
        f"Excel file has {len(workbook.sheetnames)} sheet(s): "
        f"{', '.join(workbook.sheetnames)}"
    )
    return sheets


def parse_pre_experiment_sheet(ws) -> Dict[str, object]:
    """Parse the PreExperiment sheet and return a dictionary of configuration values."""

    config: Dict[str, object] = {}
    defaults = {
        "AssayDescription": "",
        "AssayUser": "User",
        "CreationTime": datetime.now().strftime("%m-%d-%Y %H:%M:%S"),
        "ModificationTime": datetime.now().strftime("%m-%d-%Y %H:%M:%S"),
        "PreAssayShakerASpeed": 1000,
        "PreAssayShakerBSpeed": 1000,
        "PreAssayTime": 300,
        "GapTime": 200,
        "AssayShakerATemperature": 25,
        "AssayShakerBTemperature": 25,
        "bRegeneration": True,
        "bRegenerationStart": False,
        "bPlateAFlat": True,
        "AssayTemperature": 25,
    }
    config.update(defaults)

    for row in ws.iter_rows(values_only=True):
        if row[0] and isinstance(row[0], str):
            key = row[0].strip()
            value = row[1] if len(row) > 1 else None

            if "Assay Description" in key:
                config["AssayDescription"] = str(value) if value else ""
            elif "Assay User" in key:
                config["AssayUser"] = str(value) if value else "User"
            elif "Creation Time" in key:
                config["CreationTime"] = str(value) if value else config["CreationTime"]
            elif "Modification Time" in key:
                config["ModificationTime"] = str(value) if value else config["ModificationTime"]
            elif "RegenerationStart" in key:
                config["bRegenerationStart"] = bool(value) if value is not None else False
            elif "Regeneration" in key and "Start" not in key:
                config["bRegeneration"] = bool(value) if value is not None else True
            elif "PlateAFlat" in key or "Plate A Flat" in key:
                config["bPlateAFlat"] = bool(value) if value is not None else True
            elif "PreAssayTime" in key:
                config["PreAssayTime"] = int(value) if value is not None else 300
            elif "PreAssayShakerASpeed" in key:
                config["PreAssayShakerASpeed"] = int(value) if value is not None else 1000
            elif "PreAssayShakerBSpeed" in key:
                config["PreAssayShakerBSpeed"] = int(value) if value is not None else 1000
            elif "Assay Temperature" in key:
                temp = int(value) if value is not None else 25
                config["AssayShakerATemperature"] = temp
                config["AssayShakerBTemperature"] = temp
                config["AssayTemperature"] = temp

    return config


def parse_plate_layout(ws) -> Tuple[List[Dict[str, object]], List[Dict[str, object]]]:
    """
    Parse the Experiment sheet to extract 96 and max plate information.

    Simple sequential reading:
    - NinetySixInfo: Read from C14-H110 (row 14 is header, rows 15-110 are data)
    - MaxInfo: Read from Q14-V110 (row 14 is header, rows 15-110 are data)
    """

    ninety_six_info: List[Dict[str, object]] = []
    max_info: List[Dict[str, object]] = []

    for _ in range(96):
        ninety_six_info.append(
            {
                "Type": SampleType.Assay.EMPTY,
                "StringType": "empty",
                "Concentration": -1.0,
                "MolecularWeight": -1.0,
                "MolarConcentration": -1.0,
                "Information": "",
                "SampleID": "",
                "WellPosition": "",
            }
        )
        max_info.append(
            {
                "Type": SampleType.MaxPlate.EMPTY,
                "StringType": "empty",
                "Concentration": -1.0,
                "MolecularWeight": -1.0,
                "MolarConcentration": -1.0,
                "Information": "",
                "SampleID": "",
                "WellPosition": "",
            }
        )

    sample_idx = 0
    for excel_row_idx in range(15, 111):
        if sample_idx >= 96:
            break

        row = list(ws.iter_rows(min_row=excel_row_idx, max_row=excel_row_idx, values_only=True))[0]

        well_position_cell = row[1] if len(row) > 1 else None
        well_position = str(well_position_cell).strip() if well_position_cell else ""

        sample_id_cell = row[2] if len(row) > 2 else None
        sample_id = str(sample_id_cell).strip() if sample_id_cell else ""

        type_cell = row[3] if len(row) > 3 else None
        type_str = str(type_cell).strip() if type_cell is not None else ""
        sample_type = map_assay_label_to_code(type_str)

        if sample_type in (SampleType.Assay.Regeneration, SampleType.Assay.Neutralization):
            sample_id = ""

        conc_cell = row[4] if len(row) > 4 else None
        concentration = -1.0
        if conc_cell:
            try:
                concentration = float(conc_cell)
            except (ValueError, TypeError):
                pass

        mw_cell = row[5] if len(row) > 5 else None
        molecular_weight = -1.0
        if mw_cell:
            try:
                molecular_weight = float(mw_cell)
            except (ValueError, TypeError):
                pass

        m_conc_cell = row[6] if len(row) > 6 else None
        molar_concentration = -1.0
        if m_conc_cell:
            try:
                molar_concentration = float(m_conc_cell)
            except (ValueError, TypeError):
                pass
        elif concentration != -1.0 and molecular_weight not in (-1.0, 0):
            molar_concentration = (concentration / molecular_weight) * 1000

        info_cell = row[7] if len(row) > 7 else None
        information = str(info_cell).strip() if info_cell else ""

        sample_id_lower = sample_id.strip().lower()
        if (
            sample_type == SampleType.Assay.Sample
            and (not sample_id_lower or sample_id_lower in {"sample", "probe"})
            and (
                concentration in (-1.0, 0.0)
                or concentration is None
                or concentration in {"", "N/A"}
                or (isinstance(concentration, (int, float)) and concentration <= 0)
            )
        ):
            sample_type = SampleType.Assay.Buffer
            sample_id = ""
        ninety_six_info[sample_idx] = {
            "Type": sample_type,
            "StringType": type_str,
            "Concentration": concentration,
            "MolecularWeight": molecular_weight,
            "MolarConcentration": molar_concentration,
            "Information": information,
            "SampleID": sample_id,
            "WellPosition": well_position,
        }
        sample_idx += 1

    probe_idx = 0
    for excel_row_idx in range(15, 111):
        if probe_idx >= 96:
            break

        row = list(ws.iter_rows(min_row=excel_row_idx, max_row=excel_row_idx, values_only=True))[0]

        well_position_cell = row[15] if len(row) > 15 else None
        well_position = str(well_position_cell).strip() if well_position_cell else ""

        sample_id_cell = row[16] if len(row) > 16 else None
        sample_id = str(sample_id_cell).strip() if sample_id_cell else ""

        type_cell = row[17] if len(row) > 17 else None
        type_str = str(type_cell).strip() if type_cell is not None else ""

        probe_type = map_max_plate_label_to_code(type_str)
        if probe_type in {SampleType.MaxPlate.Regeneration, SampleType.MaxPlate.Neutralization}:
            sample_id = ""

        conc_cell = row[18] if len(row) > 18 else None
        concentration = -1.0
        if conc_cell and probe_type != SampleType.MaxPlate.EMPTY:
            try:
                concentration = float(conc_cell)
            except (ValueError, TypeError):
                pass

        mw_cell = row[19] if len(row) > 19 else None
        molecular_weight = -1.0
        if mw_cell and probe_type != SampleType.MaxPlate.EMPTY:
            try:
                molecular_weight = float(mw_cell)
            except (ValueError, TypeError):
                pass

        m_conc_cell = row[20] if len(row) > 20 else None
        molar_concentration = -1.0
        if m_conc_cell and probe_type != SampleType.MaxPlate.EMPTY:
            try:
                molar_concentration = float(m_conc_cell)
            except (ValueError, TypeError):
                pass
        elif (
            probe_type != SampleType.MaxPlate.EMPTY
            and concentration != -1.0
            and molecular_weight not in (-1.0, 0)
        ):
            molar_concentration = (concentration / molecular_weight) * 1000

        info_cell = row[21] if len(row) > 21 else None
        information = str(info_cell).strip() if info_cell else ""

        sample_id_lower = sample_id.strip().lower()
        if (
            probe_type == SampleType.MaxPlate.Sample
            and (not sample_id_lower or sample_id_lower in {"sample", "probe"})
            and (
                concentration in (-1.0, 0.0)
                or concentration is None
                or concentration in {"", "N/A"}
                or (isinstance(concentration, (int, float)) and concentration <= 0)
            )
        ):
            probe_type = SampleType.MaxPlate.Buffer
            sample_id = ""
        if probe_type == SampleType.MaxPlate.EMPTY:
            max_info[probe_idx] = {
                "Type": probe_type,
                "StringType": type_str,
                "Concentration": 0.0,
                "MolecularWeight": 0.0,
                "MolarConcentration": 0.0,
                "Information": None,
                "SampleID": sample_id,
                "WellPosition": well_position,
            }
        else:
            max_info[probe_idx] = {
                "Type": probe_type,
                "StringType": type_str,
                "Concentration": concentration,
                "MolecularWeight": molecular_weight,
                "MolarConcentration": molar_concentration,
                "Information": information,
                "SampleID": sample_id if sample_id else "",
                "WellPosition": well_position,
            }

        probe_idx += 1

    return ninety_six_info, max_info


def parse_assay_sheet(ws):
    """
    Parse the Assay sheet to extract regeneration settings and assay steps.

    Returns: (regeneration_settings, flow_settings, assay_steps)
    """

    rs = {
        "name": None,
        "repeat": 3,
        "RTime": 5,
        "RSpeed": 1000,
        "NTime": 5,
        "NSpeed": 1000,
    }
    fs = {
        "name": None,
        "repeatRP": 5,
        "ReaTime": 5,
        "ReaSpeed": 1000,
        "Rea2Time": 5,
        "Rea2Speed": 1000,
        "PriTime": 5,
        "PriSpeed": 1000,
        "CapTime": 120,
        "CapSpeed": 1000,
        "W1Time": 10,
        "W1Speed": 1000,
        "W2Time": 10,
        "W2Speed": 1000,
        "W3Time": 10,
        "W3Speed": 1000,
        "W4Time": 10,
        "W4Speed": 1000,
        "W5Time": 10,
        "W5Speed": 1000,
        "W6Time": 10,
        "W6Speed": 1000,
    }

    in_rs_section = False
    for row in ws.iter_rows(values_only=True):
        if row[0] and isinstance(row[0], str):
            key = str(row[0]).strip().lower()
            value = row[1] if len(row) > 1 else None

            if "regeneration step" in key or "rs" in key:
                in_rs_section = True
                continue

            if in_rs_section and value is not None:
                if "repeat" in key:
                    rs["repeat"] = int(value)
                elif "rtime" in key:
                    rs["RTime"] = int(value)
                elif "rspeed" in key:
                    rs["RSpeed"] = int(value)
                elif "ntime" in key:
                    rs["NTime"] = int(value)
                elif "nspeed" in key:
                    rs["NSpeed"] = int(value)

    steps_dict: Dict[int, List[dict]] = {}
    for row in ws.iter_rows(values_only=True):
        if (
            not row[0]
            or not isinstance(row[0], (int, float))
            or not row[1]
            or not row[2]
            or not row[3]
        ):
            continue

        # Check if row[1], row[2], row[3] are numeric before converting
        try:
            loop_num = int(row[0])
            plate = int(row[1])
            column = int(row[2])
            probe_column = int(row[3])
        except (ValueError, TypeError):
            continue

        step_type_str = str(row[4]).strip().lower() if len(row) > 4 and row[4] else ""
        try:
            time = int(row[5]) if len(row) > 5 and row[5] else 0
        except (ValueError, TypeError):
            time = 0
        try:
            speed = int(row[6]) if len(row) > 6 and row[6] else 0
        except (ValueError, TypeError):
            speed = 0

        sample_column = (plate - 1) * 12 + column

        step_type = None
        if "baseline" in step_type_str or "capture" in step_type_str:
            step_type = 0
        elif "loading" in step_type_str or "load" in step_type_str:
            step_type = 1
        elif "assoc" in step_type_str or "association" in step_type_str:
            step_type = 2
        elif "dissoc" in step_type_str or "dissociation" in step_type_str:
            step_type = 3
        else:
            continue

        step = {
            "type": step_type,
            "isMultiStep": False,
            "isSubElementExist": False,
            "listProbeColumn": [probe_column],
            "listStep": [{"SampleColumn": sample_column, "Speed": speed, "Time": time}],
        }

        steps_dict.setdefault(loop_num, []).append(step)

    steps = [steps_dict[loop_num] for loop_num in sorted(steps_dict.keys())]
    return rs, fs, steps if steps else None


def create_pre_experiment_config(pre_exp_data: Dict[str, object]) -> Dict[str, object]:
    """Create the PreExperiment configuration JSON from parsed data."""

    now = datetime.now()
    time_str = now.strftime("%m-%d-%Y %H:%M:%S")

    config = {
        "AssayDescription": pre_exp_data.get("AssayDescription", ""),
        "AssayUser": pre_exp_data.get("AssayUser", "User"),
        "CreationTime": pre_exp_data.get("CreationTime", time_str),
        "ModificationTime": pre_exp_data.get("ModificationTime", time_str),
        "StartExperimentTime": None,
        "EndExperimentTime": None,
        "AnotherSavePath": "",
        "AssayType": 2,
        "PreAssayShakerASpeed": pre_exp_data.get("PreAssayShakerASpeed", 1000),
        "PreAssayShakerBSpeed": pre_exp_data.get("PreAssayShakerBSpeed", 1000),
        "PreAssayTime": pre_exp_data.get("PreAssayTime", 300),
        "GapTime": 200,
        "AssayShakerATemperature": pre_exp_data.get("AssayShakerATemperature", 25),
        "AssayShakerBTemperature": pre_exp_data.get("AssayShakerBTemperature", 25),
        "MachineRealType": "GatorPrime",
        "idleShakerATemperature": 30,
        "idleShakerBTemperature": 30,
        "PlateAType": 0,
        "bPlateAFlat": pre_exp_data.get("bPlateAFlat", True),
        "bRegeneration": pre_exp_data.get("bRegeneration", True),
        "bRegenerationStart": pre_exp_data.get("bRegenerationStart", False),
        "RegenerationNum": 99999,
        "IsOpenRNAfterAssay": {str(i): True for i in range(7)},
        "IsOpenRNBeforeAssay": {str(i): True for i in range(7)},
        "RegenerationMode": 0,
        "ShakerASpeedDeviationLow": 0,
        "ShakerBSpeedDeviationLow": 0,
        "ShakerASpeedDeviationHigh": 0,
        "ShakerBSpeedDeviationHigh": 0,
        "ShakerATempDeviation": 0.0,
        "ShakerBTempDeviation": 0.0,
        "SpectrometerTempDeviation": 0.0,
        "ParentName": None,
        "ResultName": None,
        "SoftWareVersion": "2.17.7.0416",
        "SeriesNo": None,
        "ErrorChannelList": None,
        "ErrorPreAssayList": None,
        "ErrorSampleList": None,
        "ErrorRegenerationList": None,
        "AnalysisSettingList": [],
        "ReportSettingList": [],
        "PlateColumns": [12, 12],
        "threshold": "Infinity",
        "bsingle": False,
        "bThresh": False,
        "strFileID": None,
    }

    return config


def create_experiment_config(
    ninety_six_info: List[Dict[str, object]],
    max_info: List[Dict[str, object]],
    rs: Dict[str, object],
    fs: Dict[str, object],
    assay_steps: Optional[List[List[dict]]] = None,
) -> Dict[str, object]:
    """Create the Experiment configuration JSON."""

    if assay_steps is None:
        assay_steps = [[
            {
                "type": 0,
                "isMultiStep": False,
                "isSubElementExist": False,
                "listProbeColumn": [1],
                "listStep": [{"SampleColumn": 1, "Speed": 1000, "Time": 120}],
            },
            {
                "type": 1,
                "isMultiStep": False,
                "isSubElementExist": False,
                "listProbeColumn": [1],
                "listStep": [{"SampleColumn": 2, "Speed": 1000, "Time": 120}],
            },
            {
                "type": 0,
                "isMultiStep": False,
                "isSubElementExist": False,
                "listProbeColumn": [1],
                "listStep": [{"SampleColumn": 3, "Speed": 1000, "Time": 60}],
            },
            {
                "type": 2,
                "isMultiStep": False,
                "isSubElementExist": False,
                "listProbeColumn": [1],
                "listStep": [{"SampleColumn": 4, "Speed": 1000, "Time": 200}],
            },
            {
                "type": 3,
                "isMultiStep": False,
                "isSubElementExist": False,
                "listProbeColumn": [1],
                "listStep": [{"SampleColumn": 5, "Speed": 1000, "Time": 400}],
            },
        ]]

    config = {
        "NinetySixInfo": ninety_six_info,
        "MaxInfo": max_info,
        "rs": rs,
        "fs": fs,
        "listLoopStep": assay_steps,
        "ErrorMessage": None,
        "WarningMessage": "",
        "ConcentrationUnit": 0,
        "MolarConcentrationUnit": 0,
    }

    return config


__all__ = [
    "SampleType",
    "parse_pre_experiment_sheet",
    "parse_plate_layout",
    "parse_assay_sheet",
    "create_pre_experiment_config",
    "create_experiment_config",
]

