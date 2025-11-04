#!/usr/bin/env python3
"""Helper script to examine Excel file structure"""

import sys
try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--user", "openpyxl"])
    from openpyxl import load_workbook

if len(sys.argv) < 2:
    print("Usage: python examine_excel.py <excel_file>")
    sys.exit(1)

excel_file = sys.argv[1]
wb = load_workbook(excel_file, data_only=True)
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
print("\nFirst 30 rows:")
print("=" * 80)

for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i > 30:
        break
    print(f"Row {i:2d}: {[str(cell)[:20] if cell is not None else '' for cell in row]}")

print("\n" + "=" * 80)
print("\nSearching for key terms...")
all_values = []
for row in ws.iter_rows(values_only=True):
    for cell in row:
        if cell:
            all_values.append(str(cell).lower())

keywords = ['well', 'sample', 'concentration', 'probe', 'step', 'time', 'speed', 'plate']
for keyword in keywords:
    matches = [v for v in all_values if keyword in v]
    if matches:
        print(f"\nFound '{keyword}': {matches[:5]}")
