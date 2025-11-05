#!/usr/bin/env python3
"""
GatorBio Excel to ASY Converter

This script reads a GatorBio Assay Form Excel file and converts it to a .asy file
that can be imported into the GatorBio BLI machine software.
"""

import json
import sys
import argparse
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


class SampleType:
    """Sample type constants from the .asy file"""
    EMPTY = 0
    Sample = 1  # Analyte samples
    BACKGROUND = 2  # Background control
    Buffer = 4  # Positive control (blank)
    Load = 5  # Regenerant
    Regeneration = 6  # Negative control
    Neutralization = 7  # Reference


def map_sample_type_label_to_code(label):
    """
    Map sample type labels from Excel to numeric codes.
    Labels: Buffer, Load, Sample, Regeneration, Neutralization, Probe, Background, Negative, Reference
    """
    if not label:
        return SampleType.Buffer  # Default to empty/positive control
    
    label_lower = str(label).strip().lower()
    
    if "buffer" in label_lower or "baseline" in label_lower:
        return SampleType.Buffer  # 4
    elif "load" in label_lower:
        return SampleType.Load  # 5 - Load is typically regenerant
    elif "sample" in label_lower or "analyte" in label_lower:
        return SampleType.Sample  # 1
    elif "regeneration" in label_lower or "regen" in label_lower:
        return SampleType.Regeneration  # 6 - Regeneration
    elif "neutralization" in label_lower:
        return SampleType.Neutralization  # 7 - Neutralization
    elif "probe" in label_lower:
        return SampleType.Sample  # 1 - Probe is treated as analyte in ProbeInfo
    elif "background" in label_lower:
        return SampleType.BACKGROUND  # 2
    elif "negative" in label_lower:
        return SampleType.Regeneration  # 6
    elif "reference" in label_lower:
        return SampleType.Neutralization  # 7
    else:
        return SampleType.Buffer  # Default


def read_excel_file(excel_path):
    """Read and parse the Excel file with multiple sheets (supports .xlsx and .xlsm)"""
    wb = load_workbook(excel_path, data_only=True)
    
    # Get all sheets
    sheets = {}
    for sheet_name in wb.sheetnames:
        sheets[sheet_name] = wb[sheet_name]
    
    print(f"Excel file has {len(wb.sheetnames)} sheet(s): {', '.join(wb.sheetnames)}")
    
    return sheets


def parse_pre_experiment_sheet(ws):
    """Parse the PreExperiment sheet and return a dictionary of configuration values"""
    config = {}
    
    # Default values
    defaults = {
        "AssayDescription": "",
        "AssayUser": "User",
        "CreationTime": datetime.now().strftime("%m-%d-%Y %H:%M:%S"),
        "ModificationTime": datetime.now().strftime("%m-%d-%Y %H:%M:%S"),
        "PreAssayShakerASpeed": 1000,
        "PreAssayShakerBSpeed": 1000,
        "PreAssayTime": 300,
        "GapTime": 200,
        "AssayShakerATemperature": 25,
        "AssayShakerBTemperature": 25,
        "bRegeneration": True,
        "bRegenerationStart": False,
        "bPlateAFlat": True,
        "AssayTemperature": 25,
    }
    
    config.update(defaults)
    
    # Parse key-value pairs from the sheet
    for row in ws.iter_rows(values_only=True):
        if row[0] and isinstance(row[0], str):
            key = row[0].strip()
            value = row[1] if len(row) > 1 else None
            
            # Map Excel keys to config keys
            if "Assay Description" in key:
                config["AssayDescription"] = str(value) if value else ""
            elif "Assay User" in key:
                config["AssayUser"] = str(value) if value else "User"
            elif "Creation Time" in key:
                config["CreationTime"] = str(value) if value else config["CreationTime"]
            elif "Modification Time" in key:
                config["ModificationTime"] = str(value) if value else config["ModificationTime"]
            elif "RegenerationStart" in key:
                config["bRegenerationStart"] = bool(value) if value is not None else False
            elif "Regeneration" in key and "Start" not in key:
                config["bRegeneration"] = bool(value) if value is not None else True
            elif "PlateAFlat" in key or "Plate A Flat" in key:
                config["bPlateAFlat"] = bool(value) if value is not None else True
            elif "PreAssayTime" in key:
                config["PreAssayTime"] = int(value) if value is not None else 300
            elif "PreAssayShakerASpeed" in key:
                config["PreAssayShakerASpeed"] = int(value) if value is not None else 1000
            elif "PreAssayShakerBSpeed" in key:
                config["PreAssayShakerBSpeed"] = int(value) if value is not None else 1000
            elif "Assay Temperature" in key:
                temp = int(value) if value is not None else 25
                config["AssayShakerATemperature"] = temp
                config["AssayShakerBTemperature"] = temp
                config["AssayTemperature"] = temp
    
    return config


def parse_plate_layout(ws):
    """
    Parse the Experiment sheet to extract sample and probe information.
    Simple sequential reading:
    - SampleInfo: Read from C14-H110 (row 14 is header, rows 15-110 are data)
    - ProbeInfo: Read from Q14-V110 (row 14 is header, rows 15-110 are data)
    """
    samples = []
    probe_info = []
    
    # Initialize arrays for 96 wells with defaults
    for i in range(96):
        samples.append({
            "Type": SampleType.Buffer,
            "Concentration": -1.0,
            "MolecularWeight": -1.0,
            "MolarConcentration": -1.0,
            "Information": "",
            "SampleID": ""
        })
        probe_info.append({
            "Type": SampleType.Sample,
            "Concentration": -1.0,
            "MolecularWeight": -1.0,
            "MolarConcentration": -1.0,
            "Information": "",
            "SampleID": ""
        })
    
    # Parse SampleInfo from C14-H110 (columns C, D, E, F, G, H)
    # Row 14 is header, rows 15-110 are data rows
    # Columns: C=SampleID, D=Type, E=Concentration, F=MW, G=M Conc, H=Information
    sample_idx = 0
    for excel_row_idx in range(15, 111):  # Rows 15-110
        if sample_idx >= 96:
            break
        
        row = list(ws.iter_rows(min_row=excel_row_idx, max_row=excel_row_idx, values_only=True))[0]
        
        # Column C (index 2): SampleID
        sample_id_cell = row[2] if len(row) > 2 else None
        sample_id = str(sample_id_cell).strip() if sample_id_cell else ""
        
        # Column D (index 3): Type
        type_cell = row[3] if len(row) > 3 else None
        type_str = str(type_cell).strip() if type_cell else ""
        sample_type = map_sample_type_label_to_code(type_str)
        
        # For types 6 (Regeneration) and 7 (Neutralization), set SampleID to "N/A"
        if sample_type == SampleType.Regeneration or sample_type == SampleType.Neutralization:
            sample_id = "N/A"
        
        # Column E (index 4): Concentration
        conc_cell = row[4] if len(row) > 4 else None
        concentration = -1.0
        if conc_cell:
            try:
                concentration = float(conc_cell)
            except (ValueError, TypeError):
                pass
        
        # Column F (index 5): Molecular Weight
        mw_cell = row[5] if len(row) > 5 else None
        molecular_weight = -1.0
        if mw_cell:
            try:
                molecular_weight = float(mw_cell)
            except (ValueError, TypeError):
                pass
        
        # Column G (index 6): Molar Concentration
        m_conc_cell = row[6] if len(row) > 6 else None
        molar_concentration = -1.0
        if m_conc_cell:
            try:
                molar_concentration = float(m_conc_cell)
            except (ValueError, TypeError):
                pass
        elif concentration != -1.0 and molecular_weight != -1.0 and molecular_weight > 0:
            # Calculate molar concentration if not provided
            molar_concentration = (concentration / molecular_weight) * 1000  # Convert to µM
        
        # Column H (index 7): Information
        info_cell = row[7] if len(row) > 7 else None
        information = str(info_cell).strip() if info_cell else ""
        
        # Update SampleInfo
        samples[sample_idx] = {
            "Type": sample_type,
            "Concentration": concentration,
            "MolecularWeight": molecular_weight,
            "MolarConcentration": molar_concentration,
            "Information": information,
            "SampleID": sample_id
        }
        
        sample_idx += 1
    
    # Parse ProbeInfo from Q14-V110 (columns Q, R, S, T, U, V)
    # Row 14 is header, rows 15-110 are data rows
    # Columns: Q=SampleID, R=Type, S=Concentration, T=MW, U=M Conc, V=Information
    probe_idx = 0
    for excel_row_idx in range(15, 111):  # Rows 15-110
        if probe_idx >= 96:
            break
        
        row = list(ws.iter_rows(min_row=excel_row_idx, max_row=excel_row_idx, values_only=True))[0]
        
        # Column Q (index 16): SampleID
        sample_id_cell = row[16] if len(row) > 16 else None
        sample_id = str(sample_id_cell).strip() if sample_id_cell else ""
        
        # Column R (index 17): Type
        type_cell = row[17] if len(row) > 17 else None
        type_str = str(type_cell).strip() if type_cell else ""
        
        # Map type string to ProbeInfo type
        probe_type = SampleType.EMPTY
        if not type_str:
            probe_type = SampleType.EMPTY  # Type 0
            sample_id = None
        elif "Probe" in type_str or "probe" in type_str:
            probe_type = SampleType.Sample  # Type 1 (Probe)
            if not sample_id or sample_id == "":
                sample_id = "Probe"
        elif "Buffer" in type_str or "buffer" in type_str:
            probe_type = SampleType.Load  # Type 5 (Buffer)
            if sample_id == "Buffer":
                sample_id = ""
        elif "Sample" in type_str or "sample" in type_str:
            probe_type = SampleType.BACKGROUND  # Type 2 (Sample)
            if sample_id == "Sample":
                sample_id = ""
        elif "Load" in type_str or "load" in type_str:
            probe_type = SampleType.Load  # Type 5
            if sample_id == "Load":
                sample_id = ""
        elif "Regeneration" in type_str or "Regen" in type_str or "regen" in type_str:
            probe_type = SampleType.Regeneration  # Type 6
            sample_id = "N/A"
        elif "Neutralization" in type_str or "neutralization" in type_str:
            probe_type = SampleType.Neutralization  # Type 7
            sample_id = "N/A"
        else:
            probe_type = SampleType.EMPTY  # Type 0
            sample_id = None
        
        # Column S (index 18): Concentration
        conc_cell = row[18] if len(row) > 18 else None
        concentration = -1.0
        if conc_cell and probe_type != SampleType.EMPTY:
            try:
                concentration = float(conc_cell)
            except (ValueError, TypeError):
                pass
        
        # Column T (index 19): Molecular Weight
        mw_cell = row[19] if len(row) > 19 else None
        molecular_weight = -1.0
        if mw_cell and probe_type != SampleType.EMPTY:
            try:
                molecular_weight = float(mw_cell)
            except (ValueError, TypeError):
                pass
        
        # Column U (index 20): Molar Concentration
        m_conc_cell = row[20] if len(row) > 20 else None
        molar_concentration = -1.0
        if m_conc_cell and probe_type != SampleType.EMPTY:
            try:
                molar_concentration = float(m_conc_cell)
            except (ValueError, TypeError):
                pass
        elif concentration != -1.0 and molecular_weight != -1.0 and molecular_weight > 0:
            # Calculate molar concentration if not provided
            molar_concentration = (concentration / molecular_weight) * 1000  # Convert to µM
        
        # Column V (index 21): Information
        info_cell = row[21] if len(row) > 21 else None
        information = str(info_cell).strip() if info_cell else ""
        
        # Update ProbeInfo
        if probe_type == SampleType.EMPTY:
            probe_info[probe_idx] = {
                "Type": probe_type,
                "Concentration": 0.0,
                "MolecularWeight": 0.0,
                "MolarConcentration": 0.0,
                "Information": None,
                "SampleID": sample_id
            }
        else:
            probe_info[probe_idx] = {
                "Type": probe_type,
                "Concentration": concentration,
                "MolecularWeight": molecular_weight,
                "MolarConcentration": molar_concentration,
                "Information": information,
                "SampleID": sample_id if sample_id else ""
            }
        
        probe_idx += 1
    
    return samples, probe_info


def parse_assay_sheet(ws):
    """
    Parse the Assay sheet to extract regeneration settings and assay steps.
    Returns: (regeneration_settings, assay_steps)
    """
    # Default regeneration settings
    rs = {
        "name": None,
        "repeat": 3,
        "RTime": 5,
        "RSpeed": 1000,
        "NTime": 5,
        "NSpeed": 1000
    }
    
    # Default flow settings
    fs = {
        "name": None,
        "repeatRP": 5,
        "ReaTime": 5,
        "ReaSpeed": 1000,
        "Rea2Time": 5,
        "Rea2Speed": 1000,
        "PriTime": 5,
        "PriSpeed": 1000,
        "CapTime": 120,
        "CapSpeed": 1000,
        "W1Time": 10,
        "W1Speed": 1000,
        "W2Time": 10,
        "W2Speed": 1000,
        "W3Time": 10,
        "W3Speed": 1000,
        "W4Time": 10,
        "W4Speed": 1000,
        "W5Time": 10,
        "W5Speed": 1000,
        "W6Time": 10,
        "W6Speed": 1000
    }
    
    # Parse regeneration settings
    in_rs_section = False
    for row in ws.iter_rows(values_only=True):
        if row[0] and isinstance(row[0], str):
            key = str(row[0]).strip().lower()
            value = row[1] if len(row) > 1 else None
            
            if "regeneration step" in key or "rs" in key:
                in_rs_section = True
                continue
            
            if in_rs_section and value is not None:
                if "repeat" in key:
                    try:
                        rs["repeat"] = int(float(value))
                    except (ValueError, TypeError):
                        pass
                elif "rtime" in key:
                    try:
                        rs["RTime"] = int(float(value))
                    except (ValueError, TypeError):
                        pass
                elif "rspeed" in key:
                    try:
                        rs["RSpeed"] = int(float(value))
                    except (ValueError, TypeError):
                        pass
                elif "ntime" in key:
                    try:
                        rs["NTime"] = int(float(value))
                    except (ValueError, TypeError):
                        pass
                elif "nspeed" in key:
                    try:
                        rs["NSpeed"] = int(float(value))
                    except (ValueError, TypeError):
                        pass
    
    # Parse assay steps
    # Steps are now grouped by a "Loop" column (user-defined loops)
    steps_dict = {}  # Dictionary to store steps by loop number
    in_steps_section = False
    header_found = False
    loop_col_idx = None
    
    for row in ws.iter_rows(values_only=True):
        # Check if this is the header row with "Loop" and "StepType"
        row_str = ' '.join([str(cell).lower() if cell else '' for cell in row[:10]])
        if "loop" in row_str and "steptype" in row_str:
            header_found = True
            in_steps_section = True
            # Find the Loop column index (should be column A, index 0)
            loop_col_idx = 0
            continue
        
        if in_steps_section and header_found and loop_col_idx is not None:
            # Parse step row: Loop, Probe (listProbeColumn), Plate, Column, Sec, rpm, StepType
            try:
                loop_num = int(float(row[loop_col_idx])) if len(row) > loop_col_idx and row[loop_col_idx] is not None else None
                probe_column = int(float(row[1])) if len(row) > 1 and row[1] else 1  # Probe column (listProbeColumn)
                plate = int(float(row[2])) if len(row) > 2 and row[2] else 1  # Plate number
                column = int(float(row[3])) if len(row) > 3 and row[3] else 1  # Column number within plate
                time = int(float(str(row[4]))) if len(row) > 4 and row[4] else 120  # Sec (shifted from index 3)
                speed = int(float(row[5])) if len(row) > 5 and row[5] else 1000  # rpm (shifted from index 4)
                step_type_str = str(row[6]).strip().lower() if len(row) > 6 and row[6] else ""  # StepType (shifted from index 5)
                
                # Skip if no loop number
                if loop_num is None:
                    continue
                
                # Skip regeneration step types
                if "regen" in step_type_str or "regeneration" in step_type_str:
                    continue
                
                # Calculate SampleColumn from plate and column
                # Plate 1 uses columns 1-12, Plate 2 uses columns 13-24
                # Formula: SampleColumn = (plate - 1) * 12 + column
                sample_column = (plate - 1) * 12 + column
            except (ValueError, TypeError):
                continue
            
            # Map step type string to numeric code
            # Note: Based on reference file, "Assoc." maps to type 2 (Dissociation)
            # and "Dissoc." maps to type 3 (Regeneration) in the loop structure
            step_type = None
            if "baseline" in step_type_str or "capture" in step_type_str:
                step_type = 0
            elif "loading" in step_type_str or "load" in step_type_str:
                step_type = 1  # Loading is like association
            elif "assoc" in step_type_str or "association" in step_type_str:
                # In the reference file, "Assoc." steps after baseline are treated as Dissociation (type=2)
                step_type = 2
            elif "dissoc" in step_type_str or "dissociation" in step_type_str:
                # In the reference file, "Dissoc." steps are treated as Regeneration (type=3)
                step_type = 3
            else:
                continue  # Skip unknown step types
            
            # Create step
            step = {
                "type": step_type,
                "isMultiStep": False,
                "isSubElementExist": False,
                "listProbeColumn": [probe_column],  # Use Probe column from Excel
                "listStep": [{"SampleColumn": sample_column, "Speed": speed, "Time": time}]
            }
            
            # Group steps by loop number (user-defined loops)
            if loop_num not in steps_dict:
                steps_dict[loop_num] = []
            steps_dict[loop_num].append(step)
    
    # Convert dictionary to list of loops, sorted by loop number
    steps = [steps_dict[loop_num] for loop_num in sorted(steps_dict.keys())]
    
    return rs, fs, steps if steps else None


def create_pre_experiment_config(pre_exp_data):
    """Create the PreExperiment configuration JSON from parsed data"""
    now = datetime.now()
    time_str = now.strftime("%m-%d-%Y %H:%M:%S")
    
    config = {
        "AssayDescription": pre_exp_data.get("AssayDescription", ""),
        "AssayUser": pre_exp_data.get("AssayUser", "User"),
        "CreationTime": pre_exp_data.get("CreationTime", time_str),
        "ModificationTime": pre_exp_data.get("ModificationTime", time_str),
        "StartExperimentTime": None,
        "EndExperimentTime": None,
        "AnotherSavePath": "",
        "AssayType": 2,
        "PreAssayShakerASpeed": pre_exp_data.get("PreAssayShakerASpeed", 1000),
        "PreAssayShakerBSpeed": pre_exp_data.get("PreAssayShakerBSpeed", 1000),
        "PreAssayTime": pre_exp_data.get("PreAssayTime", 300),
        "GapTime": 200,
        "AssayShakerATemperature": pre_exp_data.get("AssayShakerATemperature", 25),
        "AssayShakerBTemperature": pre_exp_data.get("AssayShakerBTemperature", 25),
        "MachineRealType": "GatorPrime",
        "idleShakerATemperature": 30,
        "idleShakerBTemperature": 30,
        "PlateAType": 0,
        "bPlateAFlat": pre_exp_data.get("bPlateAFlat", True),
        "bRegeneration": pre_exp_data.get("bRegeneration", True),
        "bRegenerationStart": pre_exp_data.get("bRegenerationStart", False),
        "RegenerationNum": 99999,
        "IsOpenRNAfterAssay": {str(i): True for i in range(7)},
        "IsOpenRNBeforeAssay": {str(i): True for i in range(7)},
        "RegenerationMode": 0,
        "ShakerASpeedDeviationLow": 0,
        "ShakerBSpeedDeviationLow": 0,
        "ShakerASpeedDeviationHigh": 0,
        "ShakerBSpeedDeviationHigh": 0,
        "ShakerATempDeviation": 0.0,
        "ShakerBTempDeviation": 0.0,
        "SpectrometerTempDeviation": 0.0,
        "ParentName": None,
        "ResultName": None,
        "SoftWareVersion": "2.17.7.0416",
        "SeriesNo": None,
        "ErrorChannelList": None,
        "ErrorPreAssayList": None,
        "ErrorSampleList": None,
        "ErrorRegenerationList": None,
        "AnalysisSettingList": [],
        "ReportSettingList": [],
        "PlateColumns": [12, 12],
        "threshold": "Infinity",
        "bsingle": False,
        "bThresh": False,
        "strFileID": None
    }
    
    return config


def create_experiment_config(samples, probe_info, rs, fs, assay_steps=None):
    """Create the Experiment configuration JSON"""
    
    # Default assay steps if none provided
    if assay_steps is None:
        assay_steps = [[
            {
                "type": 0,  # Capture
                "isMultiStep": False,
                "isSubElementExist": False,
                "listProbeColumn": [1],
                "listStep": [{"SampleColumn": 1, "Speed": 1000, "Time": 120}]
            },
            {
                "type": 1,  # Association
                "isMultiStep": False,
                "isSubElementExist": False,
                "listProbeColumn": [1],
                "listStep": [{"SampleColumn": 2, "Speed": 1000, "Time": 120}]
            },
            {
                "type": 0,  # Baseline
                "isMultiStep": False,
                "isSubElementExist": False,
                "listProbeColumn": [1],
                "listStep": [{"SampleColumn": 3, "Speed": 1000, "Time": 60}]
            },
            {
                "type": 2,  # Dissociation
                "isMultiStep": False,
                "isSubElementExist": False,
                "listProbeColumn": [1],
                "listStep": [{"SampleColumn": 4, "Speed": 1000, "Time": 200}]
            },
            {
                "type": 3,  # Regeneration
                "isMultiStep": False,
                "isSubElementExist": False,
                "listProbeColumn": [1],
                "listStep": [{"SampleColumn": 5, "Speed": 1000, "Time": 400}]
            }
        ]]
    
    config = {
        "SampleInfo": samples,
        "ProbeInfo": probe_info,
        "rs": rs,
        "fs": fs,
        "listLoopStep": assay_steps,
        "ErrorMessage": None,
        "WarningMessage": "",
        "ConcentrationUnit": 0,
        "MolarConcentrationUnit": 0
    }
    
    return config


def generate_asy_file(excel_path, output_path=None):
    """Main function to convert Excel file to .asy file"""
    excel_path = Path(excel_path)
    
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    # Read Excel file
    print(f"Reading Excel file: {excel_path}")
    sheets = read_excel_file(excel_path)
    
    # Parse PreExperiment sheet
    pre_exp_config_data = {}
    if "PreExperiment" in sheets:
        print("Parsing PreExperiment sheet...")
        pre_exp_config_data = parse_pre_experiment_sheet(sheets["PreExperiment"])
    else:
        print("Warning: PreExperiment sheet not found, using defaults")
    
    # Parse Experiment sheet
    samples = []
    probe_info = []
    if "Experiment" in sheets:
        print("Parsing Experiment sheet...")
        samples, probe_info = parse_plate_layout(sheets["Experiment"])
        print(f"Parsed {len([s for s in samples if s['SampleID']])} samples")
    else:
        print("Warning: Experiment sheet not found, creating empty samples")
        samples = [{"Type": 4, "Concentration": -1.0, "MolecularWeight": -1.0, 
                   "MolarConcentration": -1.0, "Information": "", "SampleID": ""} for _ in range(96)]
        probe_info = [{"Type": 1, "Concentration": -1.0, "MolecularWeight": -1.0, 
                      "MolarConcentration": -1.0, "Information": "", "SampleID": ""} for _ in range(96)]
    
    # Parse Assay sheet
    rs = {"name": None, "repeat": 3, "RTime": 5, "RSpeed": 1000, "NTime": 5, "NSpeed": 1000}
    fs = {"name": None, "repeatRP": 5, "ReaTime": 5, "ReaSpeed": 1000, "Rea2Time": 5, "Rea2Speed": 1000,
          "PriTime": 5, "PriSpeed": 1000, "CapTime": 120, "CapSpeed": 1000,
          "W1Time": 10, "W1Speed": 1000, "W2Time": 10, "W2Speed": 1000, "W3Time": 10, "W3Speed": 1000,
          "W4Time": 10, "W4Speed": 1000, "W5Time": 10, "W5Speed": 1000, "W6Time": 10, "W6Speed": 1000}
    assay_steps = None
    
    if "Assay" in sheets:
        print("Parsing Assay sheet...")
        rs, fs, assay_steps = parse_assay_sheet(sheets["Assay"])
        if assay_steps:
            print(f"Found {len(assay_steps)} assay loop(s) with {sum(len(loop) for loop in assay_steps)} total steps")
            
        else:
            print("No assay steps found in Assay sheet, using defaults")
    else:
        print("Warning: Assay sheet not found, using defaults")
    
    # Create configurations
    pre_exp_config = create_pre_experiment_config(pre_exp_config_data)
    exp_config = create_experiment_config(samples, probe_info, rs, fs, assay_steps)
    
    # Generate output filename
    if output_path is None:
        output_path = excel_path.parent / f"{excel_path.stem}.asy"
    else:
        output_path = Path(output_path)
    
    # Write .asy file
    print(f"Generating .asy file: {output_path}")
    with open(output_path, 'w') as f:
        f.write("[BasicInformation]\n")
        f.write(f'PreExperiment = "{json.dumps(pre_exp_config, separators=(",", ":"))}"\n')
        f.write(f'Experiment = "{json.dumps(exp_config, separators=(",", ":"))}"\n')
    
    print(f"Successfully generated: {output_path}")
    return output_path


def show_file_dialog():
    """Show a GUI popup to select input Excel file and output .asy file"""
    import tkinter as tk
    from tkinter import filedialog, messagebox
    
    # Create root window (hidden)
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    root.attributes('-topmost', True)  # Bring to front
    
    # Ask for input Excel file
    excel_file = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[
            ("Excel files", "*.xlsx *.xlsm"),
            ("All files", "*.*")
        ]
    )
    
    if not excel_file:
        print("No input file selected. Exiting.")
        return None, None
    
    # Ask for output .asy file
    output_file = filedialog.asksaveasfilename(
        title="Save .asy File As",
        defaultextension=".asy",
        filetypes=[
            ("ASY files", "*.asy"),
            ("All files", "*.*")
        ],
        initialfile=Path(excel_file).stem + ".asy"
    )
    
    if not output_file:
        print("No output file selected. Exiting.")
        return None, None
    
    root.destroy()
    return excel_file, output_file


def main():
    # Check if command-line arguments are provided
    if len(sys.argv) > 1:
        # Use command-line mode
        parser = argparse.ArgumentParser(
            description="Convert GatorBio Assay Form Excel file to .asy file"
        )
        parser.add_argument(
            "excel_file",
            help="Path to the Excel file (.xlsx)"
        )
        parser.add_argument(
            "-o", "--output",
            help="Output .asy file path (default: same name as Excel file)"
        )
        
        args = parser.parse_args()
        
        try:
            generate_asy_file(
                args.excel_file,
                args.output
            )
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            import traceback
            traceback.print_exc()
            sys.exit(1)
    else:
        # Use GUI mode
        try:
            excel_file, output_file = show_file_dialog()
            
            if excel_file and output_file:
                generate_asy_file(excel_file, output_file)
                # Show success message
                import tkinter as tk
                from tkinter import messagebox
                root = tk.Tk()
                root.withdraw()
                root.attributes('-topmost', True)
                messagebox.showinfo("Success", f"Successfully generated:\n{output_file}")
                root.destroy()
        except Exception as e:
            import tkinter as tk
            from tkinter import messagebox
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
            root.destroy()
            print(f"Error: {e}", file=sys.stderr)
            import traceback
            traceback.print_exc()
            sys.exit(1)


if __name__ == "__main__":
    main()
